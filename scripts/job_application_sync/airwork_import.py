"""A-2: Air Work 求人データ取込 → LISTING (0-420) upsert + 【RL】判定.

設計準拠:
- ~/Downloads/hubspot_job_application_design_for_ai_v0.2.md §12, §13, §14, §15, §16, §17, §23.2
- docs/wbs_outputs/1.11.9_媒体CSV同期実装/Phase0_LISTING_実測.md
- docs/wbs_outputs/1.11.9_媒体CSV同期実装/Phase0_統合.md
- docs/wbs_outputs/1.11.9_媒体CSV同期実装/Phase0_アカウント情報タブ_実測.md
- 参照ベース: scripts/job_application_sync/hrhacker_import.py (A-1)

================================================================================
Phase 0b 確定事項 (2026-06-26 株式会社坂本 client_code=1453036 実測)
================================================================================
- AW出力形式は **XLSX (Excel)** であり CSV ではない (ZIP圧縮で配布される)
- 265列 / 1シート ("Sheet1") / 1行目ヘッダ
- 主要列は日本語+英語キー名 例: "求人番号(job_offer_id)"
- 列 0 求人番号           → media_job_id  (= id_airwork 投入値)
- 列 1 承認状況            → approval_status (補助)
- 列 2 掲載状況            → original_status (01/02 値)
- 列 3 クライアントコード   → client_code (= 顧客識別キー, CLI と整合性検証)
- 列 7 職種名              → job_name (= hs_name 投入値)
- 列14 仕事内容            → description (補助)
- 【RL】 マークは **AW側で運用されていない** とユーザー報告
  → RL_DETECT_COLUMN = None として disable (常に未判定扱い)
- 旧 CSV 入力経路 (load_aw_csv) は 互換性のためのみ残置 (新規顧客では未使用)

================================================================================
v0.2 §24 やってはいけない設計 (遵守ガード)
================================================================================
本実装は以下の原則を厳守する。特に AW では §24-3 / §24-4 が中核ガード:

  1. 媒体間でタイトル名寄せを行わない (id_airwork 1次キー + ログインID複合)
  2. 媒体間で本文類似度名寄せを行わない
  3. AW は XLSX 未検出を「公開終了」と判断しない (§15, §24-3)
     → HubSpot側のステータスは絶対に変更しない。本スクリプトはXLSX内の
       レコードのみ処理する。
  4. AW新規は最初から「弊社管理」にしない (§24-4)
     → 既定は管理区分=「未判定」。Phase 0b 確定で【RL】未運用のため
       現状すべての新規が「未判定」となる (RL_DETECT_COLUMN=None)。
  5. 全求人数を契約求人数として扱わない (契約求人数=HR + 公開中)
  6. その他媒体に同精度要求しない
  7. 人手3媒体リアルタイム更新前提にしない (定期バッチ前提)
================================================================================

ステータス正規化 (v0.2 §13):
  "02" 掲載中  → 公開中
  "01" 未掲載  → 公開終了 (初期実装は単純化)
  その他       → None (未知値はステータス送信しない)

CLI:
  python airwork_import.py --xlsx <path> --login-id <顧客クライアントコード>
                           [--dry-run|--actual] [--limit N] [--sheet Sheet1]
  python airwork_import.py --zip  <path> --login-id ...  (ZIP→中のXLSX自動解凍)
  python airwork_import.py --csv  <path> --login-id ...  (非推奨, 旧経路のみ)

出力:
  dry-run: 更新/作成予定 + skipリスト → JSON + サマリ
  actual:  HubSpot API実行 + 同一形式ログ
  ログ:    scripts/job_application_sync/logs/airwork_import_{login_id}_{timestamp}.json
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import tempfile
import time
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Optional

import requests
from dotenv import load_dotenv

try:
    import openpyxl  # type: ignore
except ImportError:  # pragma: no cover
    openpyxl = None  # XLSX入力時のみ必要

# ============================================================================
# 環境設定
# ============================================================================
HERE = Path(__file__).resolve().parent
REPO = HERE.parent.parent
LOG_DIR = HERE / "logs"
LOG_DIR.mkdir(exist_ok=True)

_ENV_PATH = REPO / ".env"
if _ENV_PATH.exists():
    load_dotenv(_ENV_PATH)

BASE = "https://api.hubapi.com"


def _headers() -> dict:
    token = os.environ.get("HUBSPOT_ACCESS_TOKEN", "")
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


# ============================================================================
# XLSX 列マッピング (Phase 0b 実測 2026-06-26 確定)
# ============================================================================
# Air Work XLSX 必須列 (これらが欠落 → RuntimeError = 仕様変更検知)
AW_XLSX_COLUMNS: dict[str, str] = {
    "求人番号(job_offer_id)": "media_job_id",
    "クライアントコード(client_code)": "client_code",
    "掲載状況(publish_status)": "original_status",
    "承認状況(approval_status)": "approval_status",
    "職種名(title)": "job_name",
    "仕事内容(description)": "description",
}

# Air Work XLSX 任意列 (存在すれば取り込む / 欠落しても続行)
AW_XLSX_OPTIONAL_COLUMNS: dict[str, str] = {
    "勤務地名(working_location_id_jp)": "work_location_name",
    "勤務地都道府県(working_location_prefecture)": "work_location_prefecture",
    "勤務地市区町村(working_location_city_area)": "work_location_city",
}

AW_XLSX_DEFAULT_SHEET = "Sheet1"

# 旧 CSV 列マッピング (Phase 0a 時代の暫定形, 後方互換のためのみ保持)
AW_CSV_COLUMNS: dict[str, str] = {
    "求人ID": "media_job_id",
    "求人タイトル": "job_name",
    "掲載状況": "original_status",
    "管理名": "rl_flag",
}
AW_CSV_ENCODING = "utf-8-sig"

# ----------------------------------------------------------------------------
# 【RL】検出設定 (Phase 0b 確定: AW側で未運用 → disable)
# ----------------------------------------------------------------------------
# RL_DETECT_COLUMN:
#   None        → 【RL】判定を完全 disable (常に False = 全件 未判定)
#                  Phase 0b 2026-06-26 株式会社坂本 XLSX 実測で
#                  「【RL】はまだない」とユーザー報告
#   文字列キー  → 内部キー (AW_*_COLUMNS の値側) を指定して contains/exact 判定
#                  (旧 CSV 経路で "rl_flag" を有効にしたい場合などに使用)
RL_DETECT_COLUMN: Optional[str] = None
RL_MATCH_MODE = "contains"   # "exact" or "contains"
RL_MARK = "【RL】"

# 内部キー → HubSpotプロパティ内部名
HUBSPOT_PROPERTY_MAP: dict[str, str] = {
    "media_job_id": "id_airwork",
    "job_name": "hs_name",
    "url": "url_airwork",
}

# 既存値保護対象 (人間入力プロパティ)
PROTECTED_PROPS = {"hs_name", "url_airwork"}

# 媒体固定値
MEDIA_NAME = "AirWORK"  # HubSpot enum値準拠 (F0 2026-06-26 実測: AirWORK で確定、"Air Work"ではない)

# 新規プロパティ (Phase 0 統合 §2 で「新規必須」とされた内部名)
PROP_MEDIA_ORIG_STATUS = "baitai_genjoukyou_airwork"   # 媒体原ステータス_AirWORK
PROP_HS_KYUUJIN_STATUS = "kyuujin_status"               # HubSpot求人ステータス (HR と共通)
PROP_KANRI_KUBUN = "kanri_kubun"                        # 管理区分
PROP_TORIKOMI_RIYUU = "torikomi_riyuu"                  # 取込理由
PROP_SAISHUU_CSV_BI = "saishuu_csv_kenshutsu_bi"        # 最終CSV検出日 (date)
PROP_KONKAI_CSV_FLAG = "konkai_csv_kenshutsu_flag"      # 今回CSV検出フラグ (bool)
PROP_DOUKI_FILENAME = "douki_moto_filename"             # 同期元ファイル名
PROP_LAST_SYNCED = "doukisaishuujikoku"                 # 最終同期日 (既存)
PROP_MEDIA_NAME = "shuyoushukkoubaitai"                 # 主要出稿媒体 (既存enum)
PROP_AW_LOGIN_ID = "airwork_account_login_id"           # AWアカウントログインID (顧客複合キー)


# ============================================================================
# ステータス正規化 (v0.2 §13)
# ============================================================================
def map_airwork_status(original: str) -> Optional[str]:
    """AirWork 原ステータスを HubSpot共通ステータスへ正規化.

    Args:
        original: AW の生値 ("01" or "02" 想定)

    Returns:
        "公開中" / "公開終了" or None (未知値)

    注意:
        v0.2 §13 補足: 01は「公開前/掲載停止/掲載終了」が混在する可能性あり。
        厳密化するには履歴 (過去02があったか) が必要だが、初期実装は単純化し
        "01"→"公開終了" にマップする。
    """
    if original is None:
        return None
    s = str(original).strip()
    if s == "02":
        return "公開中"
    if s == "01":
        return "公開終了"
    return None


# ============================================================================
# 【RL】判定 (Phase 0b 確定: AW側で未運用 → 既定 disable)
# ============================================================================
def has_rl_mark(row: dict, column: Optional[str] = None,
                mode: Optional[str] = None) -> bool:
    """指定カラムに【RL】が出現するかを判定.

    引数:
        column: None なら module の RL_DETECT_COLUMN を参照。
                RL_DETECT_COLUMN も None なら **常に False** (= 全件 未判定)
        mode:   "contains" or "exact" (既定 RL_MATCH_MODE)

    Phase 0b 確定の挙動:
        RL_DETECT_COLUMN = None なので, デフォルト呼び出しでは **常に False**。
        → AW新規はすべて「未判定」扱い (§24-4 中核ガードを上回る安全策)。

    判定モード:
        "contains" (既定): 値に【RL】が部分一致で含まれる
        "exact":            値が【RL】と完全一致

    ★ ガード: row.get(col) が None / 空文字なら必ず False
    """
    col = column if column is not None else RL_DETECT_COLUMN
    if col is None:
        # Phase 0b 確定: AW【RL】未運用 → 全件 未判定
        return False
    val = row.get(col)
    if not val:
        return False
    s = str(val)
    m = mode if mode is not None else RL_MATCH_MODE
    if m == "exact":
        return s.strip() == RL_MARK
    # 既定 contains
    return RL_MARK in s


# ============================================================================
# 入力読込: XLSX / ZIP / CSV
# ============================================================================
def _check_client_code_consistency(rows: list[dict], login_id: str,
                                    strict: bool = True) -> set[str]:
    """各行の client_code が CLI --login-id と一致するか検証.

    顧客ループ運用の前提:
        CLI で --login-id 1453036 を渡したのに、XLSX に他の client_code が
        混入していた場合 = 顧客取り違え事故。strict=True なら即停止。

    Returns:
        不整合 client_code の集合 (strict=False で警告用に使う場合)
    """
    if not login_id:
        return set()
    mismatched: set[str] = set()
    for r in rows:
        cc = (r.get("client_code") or "").strip()
        if cc and cc != login_id:
            mismatched.add(cc)
    if mismatched and strict:
        raise ValueError(
            f"client_code 不整合検知 (顧客取り違え事故防止): "
            f"CLI --login-id={login_id} だが XLSX に別の client_code が含まれる: "
            f"{sorted(mismatched)}\n"
            f"  → --login-id を確認するか、XLSX を該当顧客のものに差し替えてください。"
        )
    return mismatched


def load_aw_xlsx(path: str | Path, sheet: str = AW_XLSX_DEFAULT_SHEET,
                 login_id: str = "", strict_client_code: bool = True
                 ) -> list[dict]:
    """Air Work XLSX を読込, 内部キー辞書のリストを返す (Phase 0b 確定経路).

    Args:
        path:                XLSX ファイルパス
        sheet:               シート名 (既定 "Sheet1")
        login_id:            顧客 client_code = CLI --login-id
        strict_client_code:  XLSX 内に別 client_code が混在したら停止 (既定 True)

    Raises:
        RuntimeError: 期待カラムが XLSX ヘッダに存在しない (= AW仕様変更検知)
        ValueError:   client_code が CLI と不整合 (strict_client_code=True)
        ImportError:  openpyxl 未インストール
    """
    if openpyxl is None:
        raise ImportError("openpyxl が必要です (pip install openpyxl)")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise RuntimeError(
                f"Air Work XLSX シート未検出: '{sheet}' / "
                f"利用可能: {wb.sheetnames}"
            )
        ws = wb[sheet]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers_raw = next(rows_iter)
        except StopIteration:
            return []
        headers: list[str] = [
            ("" if h is None else str(h)) for h in headers_raw
        ]

        # ★ 仕様変更検知ガード: 必須列が揃わなければ即停止
        missing = [c for c in AW_XLSX_COLUMNS if c not in headers]
        if missing:
            raise RuntimeError(
                f"Air Work XLSX 仕様変更検知: 期待列が見つかりません。\n"
                f"  欠落: {missing}\n"
                f"  ヘッダ先頭10: {headers[:10]}\n"
                f"  → AW_XLSX_COLUMNS の定義を実測結果に合わせて更新してください。"
            )

        idx = {h: i for i, h in enumerate(headers)}
        rows: list[dict] = []
        for raw in rows_iter:
            if raw is None:
                continue
            # 全空行スキップ
            if all(v is None or (isinstance(v, str) and not v.strip())
                   for v in raw):
                continue

            row: dict = {}
            for xlsx_col, key in AW_XLSX_COLUMNS.items():
                v = raw[idx[xlsx_col]] if idx[xlsx_col] < len(raw) else None
                row[key] = "" if v is None else str(v).strip()
            for xlsx_col, key in AW_XLSX_OPTIONAL_COLUMNS.items():
                if xlsx_col in idx:
                    i = idx[xlsx_col]
                    v = raw[i] if i < len(raw) else None
                    row[key] = "" if v is None else str(v).strip()
                else:
                    row[key] = ""
            row["airwork_account_login_id"] = login_id
            rows.append(row)
    finally:
        wb.close()

    # ★ 顧客取り違え事故防止: client_code 整合性検証
    _check_client_code_consistency(rows, login_id, strict=strict_client_code)

    return rows


def load_aw_zip(path: str | Path, sheet: str = AW_XLSX_DEFAULT_SHEET,
                login_id: str = "", strict_client_code: bool = True
                ) -> list[dict]:
    """Air Work ZIP (中身は XLSX) を解凍して読込.

    AW の配布形態が ZIP の場合に使用。1個目の .xlsx を採用する。
    """
    with zipfile.ZipFile(str(path)) as z:
        xlsx_names = [n for n in z.namelist() if n.lower().endswith(".xlsx")]
        if not xlsx_names:
            raise ValueError(f"ZIP内にXLSXが見つかりません: {path}")
        with tempfile.TemporaryDirectory() as tmp:
            extracted = z.extract(xlsx_names[0], tmp)
            return load_aw_xlsx(
                extracted, sheet=sheet, login_id=login_id,
                strict_client_code=strict_client_code,
            )


def load_aw_csv(path: str | Path, encoding: str = AW_CSV_ENCODING,
                login_id: str = "") -> list[dict]:
    """[非推奨] Air Work CSV (旧 Phase 0a 暫定形) を読込.

    Phase 0b 確定で AW 配布は XLSX となったため, 本関数は後方互換のみ。
    新規顧客では load_aw_xlsx を使うこと。

    Raises:
        RuntimeError: 期待カラムがCSVヘッダに存在しない
    """
    rows: list[dict] = []
    with open(path, encoding=encoding, newline="") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        missing = [c for c in AW_CSV_COLUMNS if c not in headers]
        if missing:
            raise RuntimeError(
                f"Air Work CSV仕様変更検知: 期待カラムが見つかりません。\n"
                f"  欠落: {missing}\n"
                f"  CSVヘッダ: {headers}\n"
                f"  → AW_CSV_COLUMNS の定義を Phase 0b 実測結果に合わせて更新してください。"
            )
        for raw in reader:
            row: dict = {}
            for csv_col, key in AW_CSV_COLUMNS.items():
                row[key] = (raw.get(csv_col) or "").strip()
            row["airwork_account_login_id"] = login_id
            rows.append(row)
    return rows


def load_aw_input(path: str | Path, login_id: str = "",
                  sheet: str = AW_XLSX_DEFAULT_SHEET,
                  strict_client_code: bool = True) -> list[dict]:
    """拡張子で自動振り分け: .xlsx / .zip / .csv に対応."""
    p = str(path).lower()
    if p.endswith(".xlsx"):
        return load_aw_xlsx(path, sheet=sheet, login_id=login_id,
                            strict_client_code=strict_client_code)
    if p.endswith(".zip"):
        return load_aw_zip(path, sheet=sheet, login_id=login_id,
                           strict_client_code=strict_client_code)
    if p.endswith(".csv"):
        return load_aw_csv(path, login_id=login_id)
    raise ValueError(
        f"未対応の入力形式: {path} (.xlsx/.zip/.csv のいずれかが必要)"
    )


# ============================================================================
# HubSpot LISTING 検索 (id_airwork + airwork_account_login_id 複合キー)
# ============================================================================
def find_hubspot_jobs(media_job_ids: list[str], login_id: str) -> dict[str, dict]:
    """(id_airwork, airwork_account_login_id) の複合キーで既存LISTINGを検索.

    顧客ループ前提: 同一 id_airwork でも顧客が異なれば別物として扱う必要があり、
    AND 検索 (filterGroups の filters 配列内) で 2 条件を結合する。

    Returns:
        {media_job_id: {id, properties}}
    """
    result: dict[str, dict] = {}
    if not media_job_ids:
        return result
    target_props = [
        "id_airwork", "hs_name", "url_airwork",
        PROP_HS_KYUUJIN_STATUS, PROP_MEDIA_ORIG_STATUS,
        PROP_KANRI_KUBUN, PROP_AW_LOGIN_ID,
    ]
    headers = _headers()
    for i in range(0, len(media_job_ids), 100):
        chunk = [j for j in media_job_ids[i:i + 100] if j]
        if not chunk:
            continue
        # AND 条件: 同一 filters 配列内に並べることで AND になる
        body = {
            "limit": 200,
            "properties": target_props,
            "filterGroups": [{"filters": [
                {"propertyName": "id_airwork", "operator": "IN", "values": chunk},
                {"propertyName": PROP_AW_LOGIN_ID, "operator": "EQ", "value": login_id},
            ]}],
        }
        r = requests.post(f"{BASE}/crm/v3/objects/0-420/search",
                          headers=headers, json=body, timeout=30)
        r.raise_for_status()
        for o in r.json().get("results", []):
            jid = (o.get("properties") or {}).get("id_airwork")
            if jid:
                result[jid] = {"id": o["id"], "properties": o.get("properties") or {}}
        time.sleep(0.1)
    return result


# ============================================================================
# プロパティビルダ
# ============================================================================
def build_update_props(row: dict, existing_props: dict, today_iso: str,
                       now_iso: str, source_filename: str,
                       recruit_site_url: str = "") -> dict:
    """既存LISTING更新用プロパティ. 既存値保護を適用.

    更新対象 (常に更新):
      - 媒体原ステータス (生値: "01" or "02")
      - HubSpot求人ステータス (正規化結果, None なら送らない)
      - 最終CSV検出日 (今日)
      - 今回CSV検出フラグ = true
      - 最終同期日 (今)
      - 同期元ファイル名

    人間入力プロパティ (既存値あれば上書きしない / 空値で潰さない):
      - hs_name (求人名)
      - url_airwork (URL)

    ★ §24-3 ガード: 本関数は呼出側で「XLSXに存在する求人」のみに適用される前提。
                    未検出 LISTING は呼出側で除外され, 本関数は呼ばれない。

    ★ §24-4 ガード: 既存LISTINGの管理区分は変更しない (一度判定された区分を尊重)。
    """
    p: dict = {}

    # 媒体原ステータス (常に上書き = 生値を保持)
    if row.get("original_status"):
        p[PROP_MEDIA_ORIG_STATUS] = row["original_status"]

    # 正規化ステータス
    normalized = map_airwork_status(row.get("original_status", ""))
    if normalized is not None:
        p[PROP_HS_KYUUJIN_STATUS] = normalized

    # 同期管理列
    p[PROP_SAISHUU_CSV_BI] = today_iso
    p[PROP_KONKAI_CSV_FLAG] = "true"
    p[PROP_LAST_SYNCED] = now_iso
    p[PROP_DOUKI_FILENAME] = source_filename

    # 既存値保護: 人間入力プロパティは「既存空 AND 入力非空」のときのみ書く
    for src_key, hs_prop in [("job_name", "hs_name"), ("url", "url_airwork")]:
        src_val = row.get(src_key) or ""
        existing_raw = existing_props.get(hs_prop)
        existing_val = existing_raw.strip() if isinstance(existing_raw, str) else ""
        if src_val and not existing_val:
            p[hs_prop] = src_val

    # URL空欄検知 → 採用サイトURL+求人IDで補完 (既存url_airworkが空のときのみ)
    if recruit_site_url and not p.get("url_airwork"):
        existing_url = existing_props.get("url_airwork")
        existing_url = existing_url.strip() if isinstance(existing_url, str) else ""
        jid = row.get("media_job_id", "")
        if not existing_url and jid:
            p["url_airwork"] = f"{recruit_site_url.rstrip('/')}/{jid}/"

    return p


def build_create_props(row: dict, today_iso: str, now_iso: str,
                       source_filename: str, login_id: str,
                       recruit_site_url: str = "") -> dict:
    """新規LISTING作成用プロパティ.

    ★ 呼出側で original_status == "02" のみを許可している前提 (§14, §23.2)。
    01 の新規は本関数を呼ばないこと。

    recruit_site_url: AW採用サイトURL(https://arwrk.net/recruit/{slug})。
      入力CSVにURLが無い時、これ+求人IDで個別求人URLを生成する(補完)。

    管理区分判定 (§24-4 中核 + Phase 0b RL未運用):
      【RL】検出 → 弊社管理 (Phase 0b 確定で現状は到達しない経路)
      それ以外    → 未判定  (Phase 0b 確定で全AW新規が常にここ)
    """
    p: dict = {}
    jid = row.get("media_job_id", "")
    if not jid:
        return p

    # 必須: 媒体ID + 顧客ログインID (複合キー)
    p["id_airwork"] = jid
    p[PROP_AW_LOGIN_ID] = login_id

    # URL (入力にあれば採用 / 無ければ採用サイトURL+求人IDで生成 / それも無ければ空)
    if row.get("url"):
        p["url_airwork"] = row["url"]
    elif recruit_site_url:
        p["url_airwork"] = f"{recruit_site_url.rstrip('/')}/{jid}/"

    # 求人名 = LISTING 必須プロパティ (Phase 0b HR編で判明)
    p["hs_name"] = row.get("job_name") or f"Air Work求人 {jid}"

    # 媒体名 enum
    p[PROP_MEDIA_NAME] = MEDIA_NAME

    # ステータス
    if row.get("original_status"):
        p[PROP_MEDIA_ORIG_STATUS] = row["original_status"]
    normalized = map_airwork_status(row.get("original_status", ""))
    if normalized is not None:
        p[PROP_HS_KYUUJIN_STATUS] = normalized

    # 管理区分・取込理由 (★ §24-4 中核ガード + Phase 0b 確定挙動)
    if has_rl_mark(row):
        p[PROP_KANRI_KUBUN] = "弊社管理"
        p[PROP_TORIKOMI_RIYUU] = "Air Work掲載中新規検出(RL)"
    else:
        p[PROP_KANRI_KUBUN] = "未判定"
        p[PROP_TORIKOMI_RIYUU] = "Air Work掲載中新規検出"

    # 同期管理
    p[PROP_SAISHUU_CSV_BI] = today_iso
    p[PROP_KONKAI_CSV_FLAG] = "true"
    p[PROP_LAST_SYNCED] = now_iso
    p[PROP_DOUKI_FILENAME] = source_filename

    return p


# ============================================================================
# HubSpot 書込 (batch)
# ============================================================================
def batch_update(updates: list[dict]) -> tuple[int, int, list[str]]:
    """batch/update を 100件チャンクで実行. (ok, ng, errors)."""
    if not updates:
        return 0, 0, []
    headers = _headers()
    ok = ng = 0
    errors: list[str] = []
    for i in range(0, len(updates), 100):
        chunk = updates[i:i + 100]
        r = requests.post(f"{BASE}/crm/v3/objects/0-420/batch/update",
                          headers=headers, json={"inputs": chunk}, timeout=60)
        if r.status_code in (200, 207):
            j = r.json()
            ok += len(j.get("results", []))
            ne = j.get("numErrors", 0)
            ng += ne
            if ne:
                errors.append(f"chunk {i//100}: {r.text[:300]}")
        else:
            ng += len(chunk)
            errors.append(f"HTTP {r.status_code}: {r.text[:300]}")
        time.sleep(0.15)
    return ok, ng, errors


def batch_create(creates: list[dict]) -> tuple[int, int, list[str], dict]:
    """batch/create を 100件チャンクで実行. (ok, ng, errors, idmap)."""
    if not creates:
        return 0, 0, [], {}
    headers = _headers()
    ok = ng = 0
    errors: list[str] = []
    idmap: dict = {}
    for i in range(0, len(creates), 100):
        chunk = creates[i:i + 100]
        r = requests.post(f"{BASE}/crm/v3/objects/0-420/batch/create",
                          headers=headers, json={"inputs": chunk}, timeout=60)
        if r.status_code in (200, 201, 207):
            j = r.json()
            for o in j.get("results", []):
                jid = (o.get("properties") or {}).get("id_airwork")
                if jid:
                    idmap[jid] = o["id"]
            ok += len(j.get("results", []))
            ne = j.get("numErrors", 0)
            ng += ne
            if ne:
                errors.append(f"chunk {i//100}: {r.text[:300]}")
        else:
            ng += len(chunk)
            errors.append(f"HTTP {r.status_code}: {r.text[:300]}")
        time.sleep(0.15)
    return ok, ng, errors, idmap


# ============================================================================
# メイン処理
# ============================================================================
def run(input_path: str, login_id: str, dry_run: bool = True,
        limit: Optional[int] = None, sheet: str = AW_XLSX_DEFAULT_SHEET,
        strict_client_code: bool = True, recruit_site_url: str = "") -> dict:
    """A-2 メインオーケストレーション. 結果サマリ辞書を返す.

    Args:
        input_path:          AW 入力ファイルパス (.xlsx / .zip / .csv 対応)
        login_id:            顧客 client_code (airwork_account_login_id)
        dry_run:             既定 True (実行計画のみ)
        limit:               先頭N件のみ処理 (テスト用)
        sheet:               XLSX シート名 (既定 "Sheet1")
        strict_client_code:  XLSX 内で別 client_code が混在したら停止
    """
    input_path = str(input_path)
    source_filename = os.path.basename(input_path)
    today_iso = datetime.now().strftime("%Y-%m-%d")
    now_iso = datetime.now().strftime("%Y-%m-%dT%H:%M:%SZ")

    print(f"=== airwork_import (dry_run={dry_run}) ===")
    print(f"INPUT: {input_path}")
    print(f"login_id: {login_id}")

    rows = load_aw_input(input_path, login_id=login_id, sheet=sheet,
                         strict_client_code=strict_client_code)
    if limit:
        rows = rows[:limit]
    print(f"入力件数: {len(rows)}")

    # 媒体求人IDが空の行を除外
    valid_rows = [r for r in rows if r.get("media_job_id")]
    skipped_no_id = len(rows) - len(valid_rows)
    if skipped_no_id:
        print(f"WARN: 媒体求人ID欠落でスキップ: {skipped_no_id} 件")

    media_job_ids = [r["media_job_id"] for r in valid_rows]

    # 既存LISTING検索 (id_airwork + login_id の AND)
    if not dry_run or os.environ.get("HUBSPOT_ACCESS_TOKEN"):
        try:
            existing = find_hubspot_jobs(media_job_ids, login_id)
            print(f"既存LISTING: {len(existing)} 件")
        except Exception as e:
            if dry_run:
                print(f"WARN: Search API失敗 (dry-run継続): {e}")
                existing = {}
            else:
                raise
    else:
        existing = {}

    updates: list[dict] = []
    creates: list[dict] = []
    skipped_01_new: list[dict] = []   # ★ 01新規はスキップ (§14, §23.2)
    update_preview: list[dict] = []
    create_preview: list[dict] = []

    for row in valid_rows:
        jid = row["media_job_id"]
        if jid in existing:
            # 既存: 更新 (ステータス問わず値で上書き)
            props = build_update_props(row, existing[jid]["properties"],
                                       today_iso, now_iso, source_filename,
                                       recruit_site_url=recruit_site_url)
            updates.append({"id": existing[jid]["id"], "properties": props})
            update_preview.append({"id_airwork": jid,
                                   "hubspot_id": existing[jid]["id"],
                                   "properties": props})
        else:
            # 新規候補: 02 (掲載中) のみ作成
            orig = (row.get("original_status") or "").strip()
            if orig != "02":
                skipped_01_new.append({"id_airwork": jid,
                                       "original_status": orig,
                                       "reason": "01_or_unknown_new"})
                continue
            props = build_create_props(row, today_iso, now_iso,
                                       source_filename, login_id,
                                       recruit_site_url=recruit_site_url)
            creates.append({"properties": props})
            create_preview.append({"id_airwork": jid, "properties": props})

    summary = {
        "input_path": input_path,
        "login_id": login_id,
        "csv_total_rows": len(rows),   # キー名は後方互換のため csv_ を維持
        "skipped_no_id": skipped_no_id,
        "valid_rows": len(valid_rows),
        "existing_listings": len(existing),
        "updates_planned": len(updates),
        "creates_planned": len(creates),
        "skipped_01_new": len(skipped_01_new),
        "dry_run": dry_run,
    }

    print("\n--- 集計 ---")
    print(f"  更新予定: {len(updates)} 件")
    print(f"  新規作成予定 (02のみ): {len(creates)} 件")
    print(f"  01新規スキップ: {len(skipped_01_new)} 件 (§14 ガード)")

    if dry_run:
        log = {
            "summary": summary,
            "updates_preview": update_preview[:10],
            "creates_preview": create_preview[:10],
            "skipped_01_new_preview": skipped_01_new[:10],
        }
    else:
        print("\n--- 本番実行 ---")
        u_ok, u_ng, u_err = batch_update(updates)
        c_ok, c_ng, c_err, idmap = batch_create(creates)
        summary.update({
            "updates_ok": u_ok, "updates_ng": u_ng,
            "creates_ok": c_ok, "creates_ng": c_ng,
        })
        print(f"  更新: OK={u_ok} NG={u_ng}")
        print(f"  作成: OK={c_ok} NG={c_ng}")
        if u_err:
            print(f"  更新エラー: {u_err[:3]}")
        if c_err:
            print(f"  作成エラー: {c_err[:3]}")
        log = {
            "summary": summary,
            "errors": {"update": u_err, "create": c_err},
            "created_idmap": idmap,
        }

    # ログ出力 (login_id をファイル名に含める)
    ts = datetime.now().strftime("%Y%m%dT%H%M%S")
    safe_login = re.sub(r"[^0-9A-Za-z_.\-]+", "_", login_id) or "unknown"
    log_path = LOG_DIR / f"airwork_import_{safe_login}_{'dry' if dry_run else 'actual'}_{ts}.json"
    log_path.write_text(json.dumps(log, ensure_ascii=False, indent=2),
                        encoding="utf-8")
    print(f"\nログ: {log_path}")

    return summary


def run_xlsx(xlsx_path: str | Path, login_id: str, dry_run: bool = True,
             limit: Optional[int] = None,
             sheet: str = AW_XLSX_DEFAULT_SHEET,
             strict_client_code: bool = True,
             recruit_site_url: str = "") -> dict:
    """orchestrator から呼ばれる入口 (XLSX 専用ラッパ).

    既存 run() を XLSX 固定で呼ぶシンプルな委譲関数。
    並列 orchestrator (fetchers/aw_orchestrator.py) から
    loop.run_in_executor 経由で呼ばれる前提のため、副作用は run() と同等
    (HubSpot API 呼び出し / ログファイル出力 / 標準出力)。

    Args:
        xlsx_path: AW XLSX ファイルパス (.xlsx 限定)
        login_id:  顧客 client_code = airwork_account_login_id
        dry_run:   既定 True (実行計画のみ)
        limit:     先頭N件のみ (テスト用)
        sheet:     XLSX シート名
        strict_client_code: XLSX 内 client_code 不整合で停止 (既定 True)

    Returns:
        run() と同じ summary 辞書
        {input_path, login_id, csv_total_rows, valid_rows, existing_listings,
         updates_planned, creates_planned, skipped_01_new, dry_run,
         (actual時) updates_ok/ng, creates_ok/ng}
    """
    return run(
        str(xlsx_path), login_id=login_id, dry_run=dry_run,
        limit=limit, sheet=sheet, strict_client_code=strict_client_code,
        recruit_site_url=recruit_site_url,
    )


def parse_args(argv=None):
    p = argparse.ArgumentParser(
        description="Air Work 取込 (LISTING upsert + Phase 0b XLSX対応)"
    )
    g = p.add_mutually_exclusive_group(required=True)
    g.add_argument("--xlsx", help="Air Work XLSX パス (Phase 0b 確定経路)")
    g.add_argument("--zip", help="Air Work ZIP パス (中身は XLSX を自動解凍)")
    g.add_argument("--csv", help="[非推奨] 旧 Phase 0a CSV パス (後方互換のみ)")
    p.add_argument("--login-id", required=True, dest="login_id",
                   help="顧客 client_code (= airwork_account_login_id)")
    p.add_argument("--sheet", default=AW_XLSX_DEFAULT_SHEET,
                   help=f"XLSX シート名 (既定 {AW_XLSX_DEFAULT_SHEET})")
    p.add_argument("--allow-mixed-client-code", action="store_true",
                   help="XLSX に別 client_code が混在しても停止しない (非推奨)")
    g2 = p.add_mutually_exclusive_group()
    g2.add_argument("--dry-run", action="store_true", default=True,
                    help="(既定) 実行計画のみ出力")
    g2.add_argument("--actual", action="store_true",
                    help="HubSpot API を実行する")
    p.add_argument("--limit", type=int, default=None,
                   help="先頭N件のみ処理 (テスト用)")
    return p.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)
    input_path = args.xlsx or args.zip or args.csv
    run(input_path, login_id=args.login_id, dry_run=not args.actual,
        limit=args.limit, sheet=args.sheet,
        strict_client_code=not args.allow_mixed_client_code)


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    main()
